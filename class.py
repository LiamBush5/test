import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HomeBase Deal Calculator"

# Define some styles
header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")  # Light blue
result_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")  # Light green
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
bold_font = Font(bold=True)
title_font = Font(bold=True, size=14)
result_font = Font(bold=True, color="0000FF", size=12)
center_align = Alignment(horizontal='center')
right_align = Alignment(horizontal='right')

# Apply styles to cells
def style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format

# Add title
ws['A1'] = "HomeBase Deal Calculator"
style_cell(ws['A1'], font=title_font, alignment=center_align)
ws.merge_cells('A1:J1')

# Define column headers - each parameter is a column
headers = [
    "Number of Shares",
    "Share Price ($)",
    "Vesting Option",
    "Vesting Factor",
    "Board Seat",
    "Board Seat Value",
    "Liability Option",
    "WineMaster Share (%)",
    "Liability Cost",
    "TOTAL DEAL VALUE"
]

# Add column headers
for i, header in enumerate(headers):
    col = get_column_letter(i + 1)
    ws[f'{col}3'] = header
    style_cell(ws[f'{col}3'], font=bold_font, fill=header_fill, alignment=center_align, border=border)
    # Make columns wider
    ws.column_dimensions[col].width = 18

# Add input values and formulas in row 4
# Number of Shares
ws['A4'] = 150000
style_cell(ws['A4'], alignment=right_align, border=border)

# Share Price
ws['B4'] = 50
style_cell(ws['B4'], alignment=right_align, border=border)

# Vesting Option - will be a dropdown
ws['C4'] = "immediate"
style_cell(ws['C4'], alignment=center_align, border=border)

# Vesting Factor - calculated based on vesting option
ws['D4'] = "=CHOOSE(MATCH(C4,{\"immediate\",\"proRata\",\"endSecondYear\"},0),1,0.88,0.76)"
style_cell(ws['D4'], alignment=right_align, border=border)

# Board Seat - will be a dropdown
ws['E4'] = "no"
style_cell(ws['E4'], alignment=center_align, border=border)

# Board Seat Value - calculated based on board seat and vesting
ws['F4'] = "=IF(E4=\"yes\",CHOOSE(MATCH(C4,{\"immediate\",\"proRata\",\"endSecondYear\"},0),50000,250000,750000),0)"
style_cell(ws['F4'], alignment=right_align, border=border, number_format='$#,##0')

# Liability Option - will be a dropdown
ws['G4'] = "homebase"
style_cell(ws['G4'], alignment=center_align, border=border)

# WineMaster Share (%)
ws['H4'] = 0
style_cell(ws['H4'], alignment=right_align, border=border)

# Liability Cost - calculated based on liability option
ws['I4'] = "=IF(G4=\"homebase\",0,IF(G4=\"winemaster\",150000,IF(G4=\"shared\",150000*H4/100,0)))"
style_cell(ws['I4'], alignment=right_align, border=border, number_format='$#,##0')

# TOTAL DEAL VALUE - final calculation
ws['J4'] = "=(A4*B4*D4)+F4-I4"
style_cell(ws['J4'], font=result_font, fill=result_fill, alignment=right_align, border=border, number_format='$#,##0')

# Add data validation for dropdowns
# Vesting Option dropdown
vesting_dv = DataValidation(
    type="list",
    formula1='"immediate,proRata,endSecondYear"',
    allow_blank=False
)
ws.add_data_validation(vesting_dv)
vesting_dv.add(ws['C4'])

# Board Seat dropdown
board_dv = DataValidation(
    type="list",
    formula1='"yes,no"',
    allow_blank=False
)
ws.add_data_validation(board_dv)
board_dv.add(ws['E4'])

# Liability Option dropdown
liability_dv = DataValidation(
    type="list",
    formula1='"homebase,winemaster,shared"',
    allow_blank=False
)
ws.add_data_validation(liability_dv)
liability_dv.add(ws['G4'])

# Add a legend section
ws['A6'] = "LEGEND:"
style_cell(ws['A6'], font=bold_font)

# Vesting options
ws['A7'] = "Vesting Options:"
ws['B7'] = "immediate = Immediate Vesting (100%)"
ws['A8'] = ""
ws['B8'] = "proRata = Pro Rata Cliff Vesting (88%)"
ws['A9'] = ""
ws['B9'] = "endSecondYear = End of Second Year (76%)"

# Board seat values
ws['D7'] = "Board Seat Values:"
ws['E7'] = "immediate: $50,000"
ws['D8'] = ""
ws['E8'] = "proRata: $250,000"
ws['D9'] = ""
ws['E9'] = "endSecondYear: $750,000"

# Liability options
ws['G7'] = "Liability Options:"
ws['H7'] = "homebase = Goes to HomeBase ($0)"
ws['G8'] = ""
ws['H8'] = "winemaster = Stays with WineMaster ($150,000)"
ws['G9'] = ""
ws['H9'] = "shared = Shared based on percentage"

# Add calculation explanation
ws['A11'] = "Calculation:"
style_cell(ws['A11'], font=bold_font)
ws['A12'] = "TOTAL DEAL VALUE = (Number of Shares × Share Price × Vesting Factor) + Board Seat Value - Liability Cost"
ws.merge_cells('A12:J12')

# Save the workbook
filename = "HomeBase_Deal_Calculator.xlsx"
wb.save(filename)
print(f"Spreadsheet created: {filename}")
print(f"Full path: {os.path.abspath(filename)}")

# Optional: Create a pandas DataFrame to show a preview of the calculation
def calculate_deal_value(num_shares=150000, share_price=50, vesting_option="immediate",
                         board_seat="no", liability_option="homebase", liability_sharing=0):
    # Constants
    vesting_factors = {
        "immediate": 1.0,
        "proRata": 0.88,
        "endSecondYear": 0.76
    }

    board_seat_values = {
        "immediate": 50000,
        "proRata": 250000,
        "endSecondYear": 750000,
        "none": 0
    }

    liability_amount = 1000000
    liability_probability = 15

    # Calculations
    share_value = num_shares * share_price
    vesting_discount = vesting_factors[vesting_option]
    discounted_share_value = share_value * vesting_discount

    board_value = board_seat_values["none"]
    if board_seat == "yes":
        board_value = board_seat_values[vesting_option]

    liability_cost = 0
    expected_liability = liability_amount * (liability_probability / 100)
    if liability_option == "winemaster":
        liability_cost = expected_liability
    elif liability_option == "shared":
        liability_cost = expected_liability * (liability_sharing / 100)

    total_value = discounted_share_value + board_value - liability_cost

    return {
        "Share Value": f"${share_value:,.0f}",
        "Vesting Discount": vesting_discount,
        "Discounted Share Value": f"${discounted_share_value:,.0f}",
        "Board Seat Value": f"${board_value:,.0f}",
        "Liability Cost": f"${liability_cost:,.0f}",
        "Total Deal Value": f"${total_value:,.0f}"
    }

# Print a sample calculation
print("\nSample Calculation:")
result = calculate_deal_value()
for key, value in result.items():
    print(f"{key}: {value}")
